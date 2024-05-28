import tkinter as tk
from tkinter import filedialog, messagebox
import os
import fitz  # PyMuPDF
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
import shutil
import re

class PDFPage:
    def __init__(self):
        self.text = ""
        self.tables = []

class PDFDocument:
    def __init__(self):
        self.pages = []

def add_pdf_data_to_excel(pdf_document, excel_file_path, sheet_name):
    wb_copia = load_workbook(excel_file_path)

    title_font = Font(name='Arial', size=12, bold=True, color="000000")
    content_font = Font(name='Arial', size=10, color="000000")
    left_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    ws_copia = wb_copia.create_sheet(title=sheet_name)

    for col, width in zip(range(1, 4), [34.83, 35.17, 20.00]):
        ws_copia.column_dimensions[get_column_letter(col)].width = width

    row_index = 1

    for pdf_page in pdf_document.pages:
        for line in pdf_page.text.split("\n"):
            if not line.strip():
                continue
            cell = ws_copia.cell(row=row_index, column=1, value=line.strip())
            ws_copia.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=3)
            cell.font = title_font if row_index == 1 else content_font

            if row_index != 1:
                if line.startswith(("●", "•", "○")) or re.match(r"^\d+\.", line.strip()):
                    cell.font = Font(name='Arial', size=10, bold=True)
                else:
                    cell.font = content_font
            cell.alignment = left_alignment
            row_index += 1

        for table in pdf_page.tables:
            for row_data in table:
                for col, cell_value in enumerate(row_data):
                    cell = ws_copia.cell(row=row_index, column=col + 1, value=cell_value.strip())
                    cell.font = content_font
                    cell.alignment = left_alignment
                row_index += 1

    # Ajustar automáticamente el ancho de las columnas y el alto de las filas basándose en el contenido
    for col in range(1, 4):
        max_length = 0
        for cell in ws_copia[get_column_letter(col)]:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws_copia.column_dimensions[get_column_letter(col)].width = max_length + 2  # Se añade un pequeño margen
    for row in ws_copia.iter_rows():
        max_height = 0
        for cell in row:
            try:
                max_height = max(max_height, len(str(cell.value).split('\n')))
            except:
                pass
        ws_copia.row_dimensions[row[0].row].height = max_height * 14  # Ajusta el alto basándose en el número de líneas

    wb_copia.save(excel_file_path)

def add_excel_data_to_excel(src_excel_file_path, dest_excel_file_path, sheet_name):
    wb_src = load_workbook(src_excel_file_path)
    ws_src = wb_src.active

    wb_dest = load_workbook(dest_excel_file_path)
    ws_dest = wb_dest.create_sheet(title=sheet_name)

    for row in ws_src.iter_rows():
        for cell in row:
            new_cell = ws_dest.cell(row=cell.row, column=cell.column, value=cell.value)
            new_cell.font = cell.font.copy() if cell.font else Font(name='Arial', size=10)
            new_cell.border = cell.border.copy()
            new_cell.fill = cell.fill.copy()
            new_cell.number_format = cell.number_format
            new_cell.protection = cell.protection.copy()
            new_cell.alignment = cell.alignment.copy()

    # Ajustar automáticamente el ancho de las columnas basándose en el contenido
    for col in ws_dest.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws_dest.column_dimensions[col_letter].width = max_length + 2  # Se añade un pequeño margen

    wb_dest.save(dest_excel_file_path)

def process_excel(excel_file_path, excel_file_copy_path):
    try:
        # Copiar la plantilla
        shutil.copyfile(excel_file_path, excel_file_copy_path)
        
        # Eliminar texto del pie de página
        remove_footer_text(excel_file_copy_path)

        messagebox.showinfo("Éxito", f"Procesamiento de Excel completado y guardado en '{excel_file_copy_path}'")
    except Exception as e:
        messagebox.showerror("Error", f"Error al procesar Excel: {e}")

def leer_pdf(pdf_path):
    doc = fitz.open(pdf_path)
    pdf_pages = []
    for page_num in range(len(doc)):
        pdf_page = PDFPage()
        page = doc.load_page(page_num)
        pdf_page.text = page.get_text()
        pdf_pages.append(pdf_page)
    return pdf_pages

def remove_footer_text(excel_file_path):
    wb = load_workbook(excel_file_path)
    for sheet in wb:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    if is_similar_text(cell.value, "GFPI-135 V01"):
                        cell.value = None
    wb.save(excel_file_path)

def is_similar_text(text1, text2):
    text1 = re.sub(r'[^\w\s]', '', text1)
    text2 = re.sub(r'[^\w\s]', '', text2)
    min_length = min(len(text1), len(text2))
    intersection_length = len(set(text1.lower()) & set(text2.lower()))
    similarity = intersection_length / min_length if min_length > 0 else 0
    return similarity >= 0.8

def seleccionar_archivos():
     file_paths = filedialog.askopenfilenames(filetypes=[("Archivos PDF", "*.pdf"), ("Archivos Excel", "*.xlsx")])
     if file_paths:
                    for file in file_paths:
                        listbox_archivos.insert(tk.END, file)

def seleccionar_plantilla():
    global plantilla_path
    plantilla_path = filedialog.askopenfilename(filetypes=[("Archivos Excel", "*.xlsx")])
    if plantilla_path:
        label_plantilla.config(text=f"Plantilla seleccionada: {os.path.basename(plantilla_path)}")

def procesar_archivos():
    if not plantilla_path:
        messagebox.showerror("Error", "Por favor, seleccione una plantilla")
        return

    file_paths = listbox_archivos.get(0, tk.END)
    if not file_paths:
        messagebox.showerror("Error", "Por favor, seleccione al menos un archivo PDF o Excel")
        return

    pdf_file_paths = [file for file in file_paths if file.endswith(".pdf")]
    excel_file_paths = [file for file in file_paths if file.endswith(".xlsx")]

    if not pdf_file_paths and not excel_file_paths:
        messagebox.showerror("Error", "Por favor, seleccione al menos un archivo PDF o Excel")
        return

    try:
        excel_file_copy_path = f"{os.path.splitext(plantilla_path)[0]}_modificado.xlsx"
        shutil.copyfile(plantilla_path, excel_file_copy_path)

        for pdf_file_path in pdf_file_paths:
            pdf_document = PDFDocument()
            pdf_pages = leer_pdf(pdf_file_path)
            pdf_document.pages.extend(pdf_pages)
            sheet_name = os.path.splitext(os.path.basename(pdf_file_path))[0]
            add_pdf_data_to_excel(pdf_document, excel_file_copy_path, sheet_name)

        for excel_file_path in excel_file_paths:
            sheet_name = os.path.splitext(os.path.basename(excel_file_path))[0]
            add_excel_data_to_excel(excel_file_path, excel_file_copy_path, sheet_name)

        remove_footer_text(excel_file_copy_path)

        messagebox.showinfo("Éxito", f"Procesamiento completado y guardado en '{excel_file_copy_path}'")
    except Exception as e:
        messagebox.showerror("Error", f"Error al procesar archivos: {e}")
        print(f"Error al procesar archivos: {e}")

def crear_gui():
    global listbox_archivos, label_plantilla, plantilla_path
    plantilla_path = None

    root = tk.Tk()
    root.title("Procesador de PDF y Excel")
    root.geometry("400x450")

    label = tk.Label(root, text="Seleccione archivos PDF y Excel")
    label.pack(pady=10)

    boton_seleccionar_archivos = tk.Button(root, text="Seleccionar archivos", command=seleccionar_archivos)
    boton_seleccionar_archivos.pack(pady=5)

    listbox_archivos = tk.Listbox(root, width=50, height=10)
    listbox_archivos.pack(pady=10)

    label_plantilla = tk.Label(root, text="Plantilla no seleccionada")
    label_plantilla.pack(pady=10)

    boton_seleccionar_plantilla = tk.Button(root, text="Seleccionar plantilla", command=seleccionar_plantilla)
    boton_seleccionar_plantilla.pack(pady=5)

    boton_procesar_archivos = tk.Button(root, text="Procesar archivos", command=procesar_archivos)
    boton_procesar_archivos.pack(pady=5)

    root.mainloop()

def add_pdf_data_to_excel(pdf_document, excel_file_path, sheet_name):
    wb_copia = load_workbook(excel_file_path)

    title_font = Font(name='Arial', size=12, bold=True, color="000000")
    content_font = Font(name='Arial', size=10, color="000000")
    left_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    ws_copia = wb_copia.create_sheet(title=sheet_name)

    for col, width in zip(range(1, 4), [34.83, 35.17, 20.00]):
        ws_copia.column_dimensions[get_column_letter(col)].width = width

    row_index = 1

    for pdf_page in pdf_document.pages:
        for line in pdf_page.text.split("\n"):
            if not line.strip():
                continue
            cell = ws_copia.cell(row=row_index, column=1, value=line.strip())
            ws_copia.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=3)
            cell.font = title_font if row_index == 1 else content_font

            if row_index != 1:
                if line.startswith(("●", "•", "○")) or re.match(r"^\d+\.", line.strip()):
                    cell.font = Font(name='Arial', size=10, bold=True)
                else:
                    cell.font = content_font
            cell.alignment = left_alignment
            row_index += 1

        for table in pdf_page.tables:
            for row_data in table:
                for col, cell_value in enumerate(row_data):
                    cell = ws_copia.cell(row=row_index, column=col + 1, value=cell_value.strip())
                    cell.font = content_font
                    cell.alignment = left_alignment
                row_index += 1

    # Ajustar automáticamente el ancho de las columnas y el alto de las filas basándose en el contenido
    for col in range(1, 4):
        max_length = 0
        for cell in ws_copia[get_column_letter(col)]:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except Exception as e:
                pass
        ws_copia.column_dimensions[get_column_letter(col)].width = max_length + 2  # Se añade un pequeño margen
    for row in ws_copia.iter_rows():
        max_height = 0
        for cell in row:
            try:
                max_height = max(max_height, len(str(cell.value).split('\n')))
            except Exception as e:
                pass
        ws_copia.row_dimensions[row[0].row].height = max_height * 14  # Ajusta el alto basándose en el número de líneas

    wb_copia.save(excel_file_path)


def process_excel(excel_file_path, excel_file_copy_path):
    try:
        # Copiar la plantilla
        shutil.copyfile(excel_file_path, excel_file_copy_path)
        
        # Eliminar texto del pie de página
        remove_footer_text(excel_file_copy_path)

        messagebox.showinfo("Éxito", f"Procesamiento de Excel completado y guardado en '{excel_file_copy_path}'")
    except Exception as e:
        messagebox.showerror("Error", f"Error al procesar Excel: {e}")

def leer_pdf(pdf_path):
    doc = fitz.open(pdf_path)
    pdf_pages = []
    for page_num in range(len(doc)):
        pdf_page = PDFPage()
        page = doc.load_page(page_num)
        pdf_page.text = page.get_text()
        pdf_pages.append(pdf_page)
    return pdf_pages

def remove_footer_text(excel_file_path):
    wb = load_workbook(excel_file_path)
    for sheet in wb:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    if is_similar_text(cell.value, "GFPI-135 V01"):
                        cell.value = None
    wb.save(excel_file_path)

def is_similar_text(text1, text2):
    text1 = re.sub(r'[^\w\s]', '', text1)
    text2 = re.sub(r'[^\w\s]', '', text2)
    min_length = min(len(text1), len(text2))
    intersection_length = len(set(text1.lower()) & set(text2.lower()))
    similarity = intersection_length / min_length if min_length > 0 else 0
    return similarity >= 0.8

crear_gui()
